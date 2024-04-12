from openpyxl import load_workbook

fn = 'timetable.xlsx'
wb = load_workbook(fn)
ws = wb['ученики']
ws['A5'] = 'Привет мир!'
wb.save(fn)
wb.close