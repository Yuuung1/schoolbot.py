import openpyxl
from openpyxl import load_workbook
import datetime


# wb = openpyxl.open('timetable.xlsx', read_only=True)
# sheet = wb['ученики']
# print(sheet['A5'].value)
#
#
# filename = 'timetable.xlsx'
# workbook = load_workbook(filename)
# book = openpyxl.open('timetable.xlsx')
# uch = workbook['ученики']
# book.active = 0
# uch1 = book.active
#
# today = datetime.date.today()
#
# for i in range(1 ,1000):
#     if uch['A'+str(i)].value == 754343745:
#         cl = uch['B'+str(i)].value
#         print(cl)
# if today.strftime('%A') == 'Wednesday':
#     sheetcl = workbook[str(cl)]
#     less2 = sheetcl['A2'].value
# print(less2)


fn = 'timetable.xlsx'
wb = load_workbook(fn)
uch = wb['ученики']
tea = wb['учителя']
workbook = openpyxl.open('timetable.xlsx', read_only=True)
uchbook = workbook['ученики']

today = datetime.date.today()

for i in range(1, 1000):
    if uchbook['A' + str(i)].value == message.from_user.id:
        cl = uchbook['B' + str(i)].value
if today.strftime('%A') == 'Wednesday':
    sheetcl = workbook[str(cl)]
    less2 = sheetcl['A2'].value

bot.send_message(message.chat.id, f'Ваше расписание на сегодня:\nВторой урок - {less2}', reply_markup=markupbuttonmenu)